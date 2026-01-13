# Librerías estándar
import os
import time
import sqlite3
from datetime import datetime
import platform
import subprocess

# Librerías de interfaz gráfica
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import customtkinter as ctk

# Librerías para manejo de Excel
import pandas as pd
import xlwings as xw
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Variables globales para las ventanas
ventana_historial_abierta = None
ventana_justificar_abierta = None

# Variable global para el historial (si está abierto)
historial_treeview_widget = None
ventana_mensaje_abierta = None
contador_label = None  # Variable para mostrar el contador
contador_minutos = 10  # Tiempo de 10 minutos para marcar la ausencia
temporizador_ausencia = None  # Referencia para el temporizador


# Función para desvanecer la ventana
def fade_out(window, interval=50):
    try:
        alpha = 1.0
        while alpha > 0:
            alpha -= 0.05
            if not window.winfo_exists():
                return  # La ventana ya no existe
            window.attributes("-alpha", alpha)
            window.update()
            time.sleep(interval / 1000.0)
        if window.winfo_exists():
            window.destroy()
    except tk.TclError:
        pass  # La ventana fue destruida antes de terminar

def fade_in(window, interval=50):
    try:
        alpha = 0.0
        while alpha < 1.0:
            alpha += 0.05
            if not window.winfo_exists():
                return  # La ventana ya no existe
            window.attributes("-alpha", alpha)
            window.update()
            time.sleep(interval / 1000.0)
    except tk.TclError:
        pass  # La ventana fue destruida

# Función para mostrar mensajes con iconos personalizados
def mostrar_mensaje(tipo, mensaje):
    global ventana_mensaje_abierta

    if ventana_mensaje_abierta:
        ventana_mensaje_abierta.destroy()

    ventana_mensaje = tk.Toplevel(ventana)
    ventana_mensaje.geometry("350x150+150+455")
    ventana_mensaje.config(bg="#2A3D66")
    ventana_mensaje.attributes("-alpha", 0.0)

    if tipo == "exito":
        icono = "✅"
        color_fondo = "#4CAF50"
    elif tipo == "error":
        icono = "❌"
        color_fondo = "#FF6347"

    mensaje_label = f"{mensaje} {icono}"

    mensaje = tk.Label(ventana_mensaje, text=mensaje_label, font=("Helvetica", 12), fg="#FFFFFF", bg=color_fondo)
    mensaje.pack(pady=40)

    fade_in(ventana_mensaje)
    ventana_mensaje.after(1000, lambda: fade_out(ventana_mensaje))

    ventana_mensaje_abierta = ventana_mensaje

# Función para registrar la asistencia
def registrar_asistencia(event=None):
    global historial_treeview_widget, contador_minutos, contador_label, temporizador_ausencia

    codigo = entrada_carnet.get().strip()
    if not codigo:
        return

    conexion = sqlite3.connect("asistencia.db")
    cursor = conexion.cursor()

    cursor.execute("SELECT id, nombre FROM estudiantes WHERE codigo_barras = ?", (codigo,))
    resultado = cursor.fetchone()

    if resultado:
        estudiante_id, nombre = resultado
        fecha_actual = datetime.now().strftime("%Y-%m-%d")

        cursor.execute(""" 
            SELECT COUNT(*) FROM asistencias 
            WHERE estudiante_id = ? AND fecha_hora LIKE ? 
        """, (estudiante_id, f"{fecha_actual}%"))
        ya_asistio = cursor.fetchone()[0] > 0

        if ya_asistio:
            mostrar_mensaje("error", f"{nombre} ya tiene asistencia hoy.")
        else:
            fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cursor.execute("INSERT INTO asistencias (estudiante_id, fecha_hora) VALUES (?, ?)", (estudiante_id, fecha_hora))
            conexion.commit()
            mostrar_mensaje("exito", f"{nombre} ha sido registrado.")

            if historial_treeview_widget:
                actualizar_historial(historial_treeview_widget)

            cursor.execute("SELECT COUNT(*) FROM asistencias WHERE fecha_hora LIKE ?", (f"{fecha_actual}%",))
            total_asistencias_hoy = cursor.fetchone()[0]

            if total_asistencias_hoy == 1:
                if temporizador_ausencia:
                    ventana.after_cancel(temporizador_ausencia)
                contador_minutos = 10
                actualizar_contador()
                temporizador_ausencia = ventana.after(60000, temporizador_ausencia_func)

    else:
        mostrar_mensaje("error", "Estudiante no encontrado")

    entrada_carnet.delete(0, tk.END)
    conexion.close()


# Función para manejar la ausencia después de 10 minutos
def temporizador_ausencia_func():
    global contador_minutos, contador_label, temporizador_ausencia
    contador_minutos -= 1
    if contador_minutos <= 0:
        mostrar_mensaje("error", "No se registró la asistencia. Marcando ausencia.")
        fecha_actual = datetime.now().strftime("%Y-%m-%d")

        conexion = sqlite3.connect("asistencia.db")
        cursor = conexion.cursor()
        cursor.execute("SELECT id FROM estudiantes")
        todos = set([r[0] for r in cursor.fetchall()])
        cursor.execute("SELECT estudiante_id FROM asistencias WHERE fecha_hora LIKE ?", (f"{fecha_actual}%",))
        asistieron = set([r[0] for r in cursor.fetchall()])
        faltaron = todos - asistieron
        for estudiante_id in faltaron:
            cursor.execute("INSERT INTO asistencias (estudiante_id, fecha_hora) VALUES (?, ?)", (estudiante_id, f"{fecha_actual} (ausente)"))
        conexion.commit()
        conexion.close()

        if historial_treeview_widget:
            actualizar_historial(historial_treeview_widget)

        contador_minutos = 10
        actualizar_contador()
        temporizador_ausencia = None
    else:
        actualizar_contador()
        temporizador_ausencia = ventana.after(60000, temporizador_ausencia_func)

def actualizar_contador():
    global contador_label, contador_minutos
    if contador_label:
        contador_label.config(text=f"Tiempo restante: {contador_minutos} minutos")

 # Ventana de Historial
def abrir_historial():
    global historial_treeview_widget, contador_label, ventana_historial_abierta

    if ventana_historial_abierta and ventana_historial_abierta.winfo_exists():
        if ventana_historial_abierta.state() == 'iconic':  # Si está minimizada
            ventana_historial_abierta.deiconify()  # Restaurar
            ventana_historial_abierta.lift()      # Traer al frente
        else:
            ventana_historial_abierta.iconify()   # Minimizar
        return

    # Crear ventana historial si no existe
    ventana_historial = tk.Toplevel(ventana)
    ventana_historial.title("Historial de Asistencia")
    ventana_historial.geometry("700x730+600+0")
    ventana_historial.config(bg="#2A3D66")

    busqueda_label = tk.Label(ventana_historial, text="Buscar por Nombre:", font=("Helvetica", 12), fg="#FFFFFF", bg="#2A3D66")
    busqueda_label.pack(pady=5)
    entrada_busqueda = tk.Entry(ventana_historial, font=("Arial", 12), width=30)
    entrada_busqueda.pack(pady=10)

    treeview_frame = tk.Frame(ventana_historial)
    treeview_frame.pack(pady=10)

    treeview = ttk.Treeview(treeview_frame, columns=("Nombre", "Fecha de Asistencia"), show="headings", height=15)
    treeview.heading("Nombre", text="Nombre del Estudiante")
    treeview.heading("Fecha de Asistencia", text="Fecha de Asistencia")
    treeview.grid(row=0, column=0, sticky="nsew")

    treeview.config(selectmode="none")

    scrollbar = ttk.Scrollbar(treeview_frame, orient="vertical", command=treeview.yview)
    scrollbar.grid(row=0, column=1, sticky="ns")
    treeview.configure(yscrollcommand=scrollbar.set)

    historial_treeview_widget = treeview

    contador_label = tk.Label(ventana_historial, text=f"Tiempo restante: {contador_minutos} minutos", font=("Helvetica", 12), fg="#FFFFFF", bg="#2A3D66")
    contador_label.pack(pady=10)

    def actualizar_historial_con_busqueda():
        conexion = sqlite3.connect("asistencia.db")
        cursor = conexion.cursor()
        busqueda = entrada_busqueda.get().strip()
        if busqueda:
            cursor.execute(""" 
                SELECT estudiantes.nombre, asistencias.fecha_hora 
                FROM asistencias 
                JOIN estudiantes ON asistencias.estudiante_id = estudiantes.id 
                WHERE estudiantes.nombre LIKE ? 
            """, (f"%{busqueda}%",))
        else:
            cursor.execute(""" 
                SELECT estudiantes.nombre, asistencias.fecha_hora 
                FROM asistencias 
                JOIN estudiantes ON asistencias.estudiante_id = estudiantes.id 
            """)
        registros = cursor.fetchall()
        for row in treeview.get_children():
            treeview.delete(row)
        for registro in registros:
            if "(justificada)" in registro[1]:
                treeview.insert("", "end", values=(registro[0], f"{registro[1]} (JUSTIFICADA)"))
            else:
                treeview.insert("", "end", values=(registro[0], registro[1]))
        conexion.close()

    entrada_busqueda.bind("<KeyRelease>", lambda event: actualizar_historial_con_busqueda())

    actualizar_historial_con_busqueda()

    def cerrar_historial():
        global historial_treeview_widget, contador_minutos, ventana_historial_abierta
        historial_treeview_widget = None
        contador_minutos = 10
        ventana_historial_abierta = None
        ventana_historial.destroy()

    ventana_historial.protocol("WM_DELETE_WINDOW", cerrar_historial)

    boton_borrar = tk.Button(ventana_historial, text="Borrar Historial", font=("Helvetica", 12), fg="#FFFFFF", bg="#800020", relief="raised", width=20,
                             command=borrar_historial)
    boton_borrar.pack(pady=10)

    ventana_historial_abierta = ventana_historial

def actualizar_historial(treeview):
    conexion = sqlite3.connect("asistencia.db")
    cursor = conexion.cursor()
    cursor.execute("SELECT estudiantes.nombre, asistencias.fecha_hora FROM asistencias JOIN estudiantes ON asistencias.estudiante_id = estudiantes.id")
    registros = cursor.fetchall()
    for row in treeview.get_children():
        treeview.delete(row)
    for registro in registros:
        if "(justificada)" in registro[1]:
            treeview.insert("", "end", values=(registro[0], f"{registro[1]} (JUSTIFICADA)"))
        else:
            treeview.insert("", "end", values=(registro[0], registro[1]))
    conexion.close()

   # Borrar Historial
def borrar_historial():
    def confirmar_borrado():
        conexion = sqlite3.connect("asistencia.db")
        cursor = conexion.cursor()
        cursor.execute("DELETE FROM asistencias")
        conexion.commit()
        conexion.close()
        if historial_treeview_widget:
            actualizar_historial(historial_treeview_widget)
        mostrar_mensaje("exito", "Historial borrado con éxito.")
        global contador_minutos, temporizador_ausencia
        contador_minutos = 10
        if isinstance(temporizador_ausencia, (str, int)):
            ventana.after_cancel(temporizador_ausencia)
        temporizador_ausencia = None
        actualizar_contador()
        popup.destroy()

    def cancelar_borrado():
        popup.destroy()

    popup = tk.Toplevel(ventana)
    popup.title("Confirmar borrado")
    popup.geometry("350x150+150+455")
    popup.configure(bg="#2A3D66")
    popup.grab_set()  # Hace modal la ventana

    label = tk.Label(
        popup,
        text="¿Estás seguro de que deseas borrar todo el historial de asistencias?",
        wraplength=300,
        fg="white",
        bg="#2c3e50"
    )
    label.pack(pady=20, padx=20)

    botones_frame = tk.Frame(popup, bg="#2c3e50")
    botones_frame.pack(pady=10)

    btn_si = tk.Button(
        botones_frame, text="Sí", width=10,
        command=confirmar_borrado,
        bg="#27ae60",  # Verde
        fg="white",
        activebackground="#2ecc71",
        activeforeground="white"
    )
    btn_si.pack(side="left", padx=10)

    btn_no = tk.Button(
        botones_frame, text="No", width=10,
        command=cancelar_borrado,
        bg="#c0392b",  # Rojo
        fg="white",
        activebackground="#e74c3c",
        activeforeground="white"
    )
    btn_no.pack(side="left", padx=10)

 # Ventana de justificación
def justificar_inasistencia():
    global ventana_justificar_abierta, entrada_nombre, entrada_fecha  # Agregar estas dos variables globales

    if ventana_justificar_abierta and ventana_justificar_abierta.winfo_exists():
        if ventana_justificar_abierta.state() == 'iconic':  # Minimizada
            ventana_justificar_abierta.deiconify()
            ventana_justificar_abierta.lift()
        else:
            ventana_justificar_abierta.iconify()
        return

    ventana_justificar = tk.Toplevel(ventana)
    ventana_justificar.title("Justificar Inasistencia")
    ventana_justificar.geometry("600x310+0+455")
    ventana_justificar.config(bg="#2A3D66")

    tk.Label(ventana_justificar, text="Nombre del estudiante:", font=("Helvetica", 12), fg="#FFFFFF", bg="#2A3D66").pack(pady=10)
    entrada_nombre = tk.Entry(ventana_justificar, font=("Arial", 12), width=30)
    entrada_nombre.pack(pady=5)

    tk.Label(ventana_justificar, text="Fecha a justificar (YYYY-MM-DD):", font=("Helvetica", 12), fg="#FFFFFF", bg="#2A3D66").pack(pady=10)
    entrada_fecha = tk.Entry(ventana_justificar, font=("Arial", 12), width=30)
    entrada_fecha.pack(pady=5)

    # Botón para guardar la justificación, ahora sí aparece
    boton_guardar = tk.Button(ventana_justificar, text="Justificar", font=("Helvetica", 12), fg="#FFFFFF", bg="#1B5E20", relief="raised", width=20, command=guardar_justificacion)
    boton_guardar.pack(pady=20)

    # Guardar la referencia a la ventana
    ventana_justificar_abierta = ventana_justificar

    # Manejar el cierre de la ventana para limpiar la variable global
    def cerrar_ventana_justificar():
        global ventana_justificar_abierta
        ventana_justificar_abierta = None
        ventana_justificar.destroy()

    ventana_justificar.protocol("WM_DELETE_WINDOW", cerrar_ventana_justificar)

def guardar_justificacion():
    import re
    global entrada_nombre, entrada_fecha  # para acceder a las entradas creadas en justificar_inasistencia
    nombre = entrada_nombre.get().strip()
    fecha = entrada_fecha.get().strip()

    if not nombre or not fecha:
        mostrar_mensaje("error", "Todos los campos son obligatorios")
        return

    if not re.match(r'^\d{4}-\d{2}-\d{2}$', fecha):
        mostrar_mensaje("error", "Formato de fecha inválido. Use YYYY-MM-DD")
        return

    conexion = sqlite3.connect("asistencia.db")
    cursor = conexion.cursor()

    cursor.execute("SELECT id FROM estudiantes WHERE nombre = ?", (nombre,))
    estudiante = cursor.fetchone()
    if not estudiante:
        mostrar_mensaje("error", "Estudiante no encontrado.")
        conexion.close()
        return

    estudiante_id = estudiante[0]

    cursor.execute("""
        SELECT id, fecha_hora FROM asistencias 
        WHERE estudiante_id = ? AND fecha_hora LIKE ?
    """, (estudiante_id, f"{fecha}%"))
    registro = cursor.fetchone()

    if registro:
        cursor.execute("""
            UPDATE asistencias 
            SET fecha_hora = ? 
            WHERE id = ?
        """, (f"{fecha} (justificada)", registro[0]))
    else:
        cursor.execute("""
            INSERT INTO asistencias (estudiante_id, fecha_hora) 
            VALUES (?, ?)
        """, (estudiante_id, f"{fecha} (justificada)"))

    conexion.commit()
    conexion.close()
    mostrar_mensaje("exito", "Inasistencia justificada correctamente.")

    if historial_treeview_widget:
        actualizar_historial(historial_treeview_widget)

    # Exportar a excel
def exportar_a_excel():
    import sqlite3
    import pandas as pd
    import os

    conexion = sqlite3.connect("asistencia.db")
    cursor = conexion.cursor()

    # Obtener lista completa de estudiantes
    cursor.execute("SELECT nombre FROM estudiantes")
    estudiantes = [row[0] for row in cursor.fetchall()]

    # Obtener registros actuales desde la base de datos
    cursor.execute("""
        SELECT estudiantes.nombre, asistencias.fecha_hora 
        FROM asistencias 
        JOIN estudiantes ON asistencias.estudiante_id = estudiantes.id
    """)
    registros = cursor.fetchall()

    nuevos_datos = []
    for nombre, fecha_estado in registros:
        if "(ausente)" in fecha_estado.lower():
            fecha = fecha_estado.replace(" (ausente)", "")
            estado = "A"
        elif "(justificada)" in fecha_estado.lower():
            fecha = fecha_estado.replace(" (justificada)", "")
            estado = "J"
        else:
            fecha = fecha_estado
            estado = "P"
        fecha = fecha.strip().split()[0]
        nuevos_datos.append((nombre, fecha, estado))

    df_nuevos = pd.DataFrame(nuevos_datos, columns=["Nombre", "Fecha", "Estado"])

    # Leer datos anteriores del Excel si existe
    archivo_excel = "historial_asistencia.xlsx"
    if os.path.exists(archivo_excel):
        try:
            df_previo = pd.read_excel(archivo_excel)
            df_previo = df_previo.melt(id_vars=["Nombre"], var_name="Fecha", value_name="Estado")
            df_previo.dropna(inplace=True)
        except Exception as e:
            mostrar_mensaje("error", f"No se pudo leer el Excel anterior: {str(e)}")
            df_previo = pd.DataFrame(columns=["Nombre", "Fecha", "Estado"])
    else:
        df_previo = pd.DataFrame(columns=["Nombre", "Fecha", "Estado"])

    # Unir datos nuevos con los anteriores
    df_total = pd.concat([df_previo, df_nuevos], ignore_index=True)
    df_total.drop_duplicates(subset=["Nombre", "Fecha"], keep="last", inplace=True)

    # Asegurarse de que todos los estudiantes estén presentes en todas las fechas
    todas_las_fechas = sorted(df_total["Fecha"].unique())
    pares_completos = [(nombre, fecha) for nombre in estudiantes for fecha in todas_las_fechas]
    df_completo = pd.DataFrame(pares_completos, columns=["Nombre", "Fecha"])

    df_merged = pd.merge(df_completo, df_total, on=["Nombre", "Fecha"], how="left")
    df_merged["Estado"] = df_merged["Estado"].fillna("A")

    # Pivotear
    tabla_pivot = df_merged.pivot(index="Nombre", columns="Fecha", values="Estado")
    tabla_pivot.sort_index(inplace=True)
    tabla_pivot = tabla_pivot.reindex(sorted(tabla_pivot.columns), axis=1)
    tabla_pivot.reset_index(inplace=True)

    # Exportar
    try:
        tabla_pivot.to_excel(archivo_excel, index=False)
        mostrar_mensaje("exito", "Historial exportado a Excel.")
    except Exception as e:
        mostrar_mensaje("error", f"Error al exportar a Excel: {str(e)}")

    conexion.close()

 # Ventana para agregar gente
def abrir_ventana_agregar_estudiante():
    if hasattr(abrir_ventana_agregar_estudiante, "ventana") and abrir_ventana_agregar_estudiante.ventana.winfo_exists():
        abrir_ventana_agregar_estudiante.ventana.focus()
        return

    ventana = ctk.CTkToplevel()
    abrir_ventana_agregar_estudiante.ventana = ventana
    ventana.title("Agregar Estudiante")
    ventana.geometry("325x425+600+0")
    ventana.resizable(False, False)

    ctk.CTkLabel(ventana, text="Cédula:").pack(pady=(20, 5))
    entrada_id = ctk.CTkEntry(ventana)
    entrada_id.pack()

    ctk.CTkLabel(ventana, text="Nombres:").pack(pady=(20, 5))
    entrada_nombre = ctk.CTkEntry(ventana)
    entrada_nombre.pack()

    ctk.CTkLabel(ventana, text="Apellidos:").pack(pady=(20, 5))
    entrada_apellido = ctk.CTkEntry(ventana)
    entrada_apellido.pack()

    ctk.CTkLabel(ventana, text="Carnet:").pack(pady=(20, 5))
    entrada_codigo_barras = ctk.CTkEntry(ventana)
    entrada_codigo_barras.pack()

    mensaje = ctk.CTkLabel(ventana, text="", text_color="red")
    mensaje.pack(pady=10)

    def agregar_estudiante(event=None):
        id = entrada_id.get().strip()
        nombre = entrada_nombre.get().strip()
        apellido = entrada_apellido.get().strip()
        codigo_barras = entrada_codigo_barras.get().strip()

        if not (id and nombre and apellido and codigo_barras):
            mensaje.configure(text="Por favor, completa todos los campos.", text_color="red")
            return

        try:
            conexion = sqlite3.connect("asistencia.db")
            cursor = conexion.cursor()

            cursor.execute("""
                INSERT INTO estudiantes (id, nombre, apellido, codigo_barras)
                VALUES (?, ?, ?, ?)
            """, (id, nombre, apellido, codigo_barras))

            conexion.commit()
            conexion.close()

            mensaje.configure(text="Estudiante agregado correctamente.", text_color="green")

            entrada_id.delete(0, ctk.END)
            entrada_nombre.delete(0, ctk.END)
            entrada_apellido.delete(0, ctk.END)
            entrada_codigo_barras.delete(0, ctk.END)
            entrada_id.focus()

        except sqlite3.IntegrityError:
            mensaje.configure(text="Ya existe un estudiante con esa cédula o carnet.", text_color="orange")
        except Exception as e:
            mensaje.configure(text=f"Error: {str(e)}", text_color="red")

    # Vincular Enter a todas las entradas
    for entry in (entrada_id, entrada_nombre, entrada_apellido, entrada_codigo_barras):
        entry.bind("<Return>", agregar_estudiante)

    # Abrir el excel

archivo_1 = r"C:\Users\erick\Desktop\asistencia_estudiantes\historial_asistencia.xlsx"
archivo_2 = r"C:\Users\erick\Desktop\asistencia_estudiantes\Asistencias CTPP.xlsm"

def abrir_excel():
    """Abre dos archivos Excel con el visor predeterminado del sistema."""
    try:
        sistema = platform.system()

        archivos = [archivo_1, archivo_2]

        for archivo in archivos:
            if sistema == "Windows":
                os.startfile(archivo)
            elif sistema == "Darwin":  # macOS
                subprocess.run(["open", archivo], check=True)
            else:  # Linux
                subprocess.run(["xdg-open", archivo], check=True)

    except FileNotFoundError:
        print("Uno de los archivos no se encontró en la ruta especificada.")
    except subprocess.CalledProcessError:
        print("Hubo un error al intentar abrir uno de los archivos.")
    except Exception as e:
        print(f"Ocurrió un error inesperado al abrir los archivos: {e}")

# Evento de salir 
def exit():
    ventana.destroy()



# Crear ventana principal
def registrar_asistencia_evento(event=None):
    registrar_asistencia()

ventana = tk.Tk()
ventana.title("Control de Asistencia")
ventana.geometry("600x425+0+0")
ventana.config(bg="#2A3D66")
ventana.overrideredirect(True)


ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# Nombre del colegio
frame_titulo = tk.Frame(ventana, bg="#2A3D66")
frame_titulo.pack(pady=(20, 10))

etiqueta_sombra = tk.Label(frame_titulo, 
                           text="Colegio Técnico Profesional de Platanar",
                           font=("Georgia", 22, "bold"),
                           fg="#1A1A1A", bg="#2A3D66")
etiqueta_sombra.place(x=3, y=3)

etiqueta_colegio = tk.Label(frame_titulo, 
                            text="Colegio Técnico Profesional de Platanar",
                            font=("Georgia", 22, "bold"),
                            fg="#121212", bg="#FAF9F6")
etiqueta_colegio.pack()

# Campo para ingresar el carnet

# Frame del carnet
frame_carnet = ctk.CTkFrame(ventana, fg_color="#2A3D66", width=350, height=50)
# Cambia la posición con x, y (coordenadas absolutas) o relx, rely (relativas)
frame_carnet.place(x=220, y=100) 

# Etiqueta "Carnet:"
etiqueta_carnet = ctk.CTkLabel(frame_carnet, text="Carnet:", font=("Arial", 16))
etiqueta_carnet.pack(side="left", padx=(10, 10))

# Entrada del carnet
entrada_carnet = ctk.CTkEntry(frame_carnet, width=200, font=("Arial", 16))
entrada_carnet.pack(side="left")

entrada_carnet.focus()

# Evento para registrar asistencia al presionar Enter
entrada_carnet.bind("<Return>", registrar_asistencia_evento)


# Botones de pagina principal

# --- Funciones de animación suave y resalte profesional ---
def aplicar_animacion_profesional(boton, color_hover="#43A047"):
    # Guardamos el tamaño original
    original_width, original_height = 170, 40
    hover_width, hover_height = 190, 50
    pressed_width, pressed_height = 180, 45
    original_fg = boton.cget("fg_color")
    
    # Función para animar tamaño gradualmente
    def animar_tamano(target_w, target_h, steps=5):
        start_w = boton.winfo_width()
        start_h = boton.winfo_height()
        dw = (target_w - start_w) / steps
        dh = (target_h - start_h) / steps

        def step(i=0):
            if i < steps:
                boton.configure(width=int(start_w + dw*(i+1)), height=int(start_h + dh*(i+1)))
                boton.after(15, step, i+1)
        step()

    # Eventos
    def on_enter(e):
        animar_tamano(hover_width, hover_height)
        boton.configure(fg_color=color_hover, border_width=2, border_color="#FFFFFF")

    def on_leave(e):
        animar_tamano(original_width, original_height)
        boton.configure(fg_color=original_fg, border_width=0)

    def on_press(e):
        animar_tamano(pressed_width, pressed_height)

    def on_release(e):
        animar_tamano(hover_width, hover_height)

    # Binds
    boton.bind("<Enter>", on_enter)
    boton.bind("<Leave>", on_leave)
    boton.bind("<ButtonPress>", on_press)
    boton.bind("<ButtonRelease>", on_release)


# --- Frame para los botones ---
frame_botones = ctk.CTkFrame(ventana, fg_color="#2A3D66")
frame_botones.pack(side="left", fill="y")

# --- Creación de botones con animación ---
botones_info = [
    ("Registrar Asistencia", "#1B5E20", registrar_asistencia),
    ("Abrir Historial", "#0D47A1", abrir_historial),
    ("Justificar Inasistencia", "#D2691E", justificar_inasistencia),
    ("Exportar a Excel", "#4B0082", exportar_a_excel),
    ("Abrir Excel", "#169273", abrir_excel),
    ("Agregar Estudiante", "#807700", abrir_ventana_agregar_estudiante),
    ("Cerrar app", "#800020", exit)
]

for texto, color, comando in botones_info:
    btn = ctk.CTkButton(
        frame_botones,
        text=texto,
        font=("Helvetica", 15),
        fg_color=color,
        hover_color=color,
        text_color="#FFFFFF",
        corner_radius=8,
        width=170,
        height=40,
        command=comando
    )
    btn.pack(pady=4)
    aplicar_animacion_profesional(btn)

ventana.mainloop()